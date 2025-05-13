import openpyxl

# 打开 Excel 文件
workbook = openpyxl.load_workbook('1.xlsx')
sheet = workbook.active  # 获取当前活动工作表

# 遍历每行（从第2行开始，跳过表头）
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    row_num = row[0].row  # 获取当前行号
    
    # 获取第二列、第三列、第五列和第六列的单元格（索引从0开始）
    cell_col2 = row[1]  # 第二列
    cell_col3 = row[2]  # 第三列
    cell_col5 = row[4]  # 第五列
    cell_col6 = row[5]  # 第六列
    
      # 获取单元格值
    value_col2 = cell_col2.value
    value_col3 = cell_col3.value
    
    # 判断第二列和第三列是否都为空
    if (value_col2 is None or str(value_col2).strip() == '') and \
       (value_col3 is None or str(value_col3).strip() == ''):
        print(f"第{row_num}行 | 第二列和第三列均为空，跳过计算")
        continue  # 跳过当前行，不进行计算
    
    # 获取单元格值，若为空则用0替代
    value_col2 = cell_col2.value or 0
    value_col3 = cell_col3.value or 0
    value_col5 = cell_col5.value or 0
    value_col6 = cell_col6.value or 0
    
    # 转换为数值类型
    try:
        num_col2 = float(value_col2)
        num_col3 = float(value_col3)
        num_col5 = float(value_col5)
        num_col6 = float(value_col6)
    except (ValueError, TypeError):
        # 若无法转换为数值，默认视为0
        num_col2 = 0
        num_col3 = 0
        num_col5 = 0
        num_col6 = 0
    
    # 计算第二列和第三列的和
    total = num_col2 + num_col3
    
    # 根据和的正负执行不同计算
    if total >= 0:
        # 正数逻辑：(1 - (第二列/100)) * (第五列/第六列)
        formula = f"{num_col2}*(1 - ({num_col5}/100)) * {num_col6}*7.2"
        if num_col6 != 0:  # 避免除以零
            result = num_col2 * (1 - num_col5/ 100) / num_col6 * 7.2
        else:
            result = 0
            formula += " = 0 (避免除以零)"
    else:
        # 负数逻辑：(1 + (第二列/100)) * (第五列/第六列)
        formula = f"{num_col3}*(1 + ({num_col5}/100)) */{num_col6}*7.2"
        if num_col6 != 0:  # 避免除以零
            result = num_col3 * (1 +num_col5/ 100) / num_col6 * 7.2
        else:
            result = 0
            formula += " = 0 (避免除以零)"
    
    # 将结果写入第七列（索引6）
    result_cell = row[7]
    result_cell.value = result
    
    # 打印公式和结果
    print(f"第{row_num}行 | 公式: {formula} = {result}")

# 保存修改
workbook.save('1.xlsx')
print(f"\n已完成计算并保存到新文件：example_with_calculations.xlsx")
