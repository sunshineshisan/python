import openpyxl

# 打开 Excel 文件
workbook = openpyxl.load_workbook('G:/x/1.xlsx')

# 选择工作表
sheet = workbook.active

# 假设我们要计算 A 列和 B 列对应单元格的和，并将结果写入 C 列
for row in range(1, sheet.max_row + 1):
    # 获取 A 列和 B 列的值
    cell_a = sheet.cell(row=row, column=2)
    cell_b = sheet.cell(row=row, column=3)
    print(cell_a,cell_b)
    # 确保单元格中有值且为数字类型
    if cell_a.value and cell_b.value and isinstance(cell_a.value, (int, float)) and isinstance(cell_b.value, (int, float)):
        # 计算和
        result = cell_a.value + cell_b.value
        
        # 将结果写入 C 列
        cell_c = sheet.cell(row=row, column=7)
        cell_c.value = result

# 保存修改后的 Excel 文件
workbook.save('G:/x/2.xlsx')
    