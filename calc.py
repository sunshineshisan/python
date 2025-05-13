import openpyxl
import os

file_path = '1.xlsx'
output_path = '1.xlsx'  # 输出文件路径

try:
    # 检查文件是否存在且可读
    if not os.path.exists(file_path):
        print(f"错误：文件 '{file_path}' 不存在")
    elif not os.access(file_path, os.R_OK):
        print(f"错误：没有读取文件 '{file_path}' 的权限")
    else:
        # 打开工作簿
        workbook = openpyxl.load_workbook(file_path)
        
        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\n正在处理工作表：'{sheet_name}'")
            
            # 遍历每行（从第2行开始，跳过表头）
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                row_num = row[0].row  # 获取当前行号
                
                # 获取第二列、第三列、第五列和第六列的单元格
                cell_col2 = row[1]  # 第二列
                cell_col3 = row[2]  # 第三列
                cell_col5 = row[4]  # 第五列
                cell_col6 = row[5]  # 第六列
                
                # 获取单元格值
                value_col2 = cell_col2.value
                value_col3 = cell_col3.value
                
                # 判断第二列和第三列是否都为空
                if (value_col2 is None or str(value_col2).strip() == '') and \
                   (value_col3 is None or str(value_col3).strip() == ''):
                    print(f"  第{row_num}行 | 第二列和第三列均为空，跳过计算")
                    continue  # 跳过当前行
                
                # 处理空值为0
                value_col2 = value_col2 or 0
                value_col3 = value_col3 or 0
                value_col5 = cell_col5.value or 0
                value_col6 = cell_col6.value or 0
                
                # 转换为数值类型
                try:
                    num_col2 = float(value_col2)
                    num_col3 = float(value_col3)
                    num_col5 = float(value_col5)
                    num_col6 = float(value_col6)
                except (ValueError, TypeError):
                    num_col2 = 0
                    num_col3 = 0
                    num_col5 = 0
                    num_col6 = 0
                
                # 计算第二列和第三列的和
                total = num_col2 + num_col3
                
                # 根据和的正负执行不同计算
                if total >= 0:
                    formula = f"{num_col2}*(1 - ({num_col5}/100)) * {num_col6}*7.2"
                    if num_col6 != 0:
                        result = num_col2 * (1 - num_col5/ 100) / num_col6 * 7.2
                    else:
                        result = 0
                        formula += " = 0 (避免除以零)"
                else:
                    formula = f"{num_col3}*(1 + ({num_col5}/100)) */{num_col6}*7.2"
                    if num_col6 != 0:
                        result = num_col3 * (1 +num_col5/ 100) / num_col6 * 7.2
                    else:
                        result = 0
                        formula += " = 0 (避免除以零)"
                
                # 将结果写入第七列
                result_cell = row[6]
                result_cell.value = result
                
                # 打印公式和结果
                print(f"  第{row_num}行 | 公式: {formula} = {result}")
        
        # 保存所有修改
        try:
            workbook.save(output_path)
            print(f"\n✅ 已完成所有工作表的计算，并保存到：{output_path}")
        except PermissionError:
            print(f"错误：没有写入文件 '{output_path}' 的权限，请检查目录权限")

except PermissionError:
    print(f"错误：没有权限访问文件 '{file_path}'")
except Exception as e:
    print(f"发生未知错误：{e}")
