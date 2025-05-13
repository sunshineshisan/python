import openpyxl

# 打开当前 XLSX 文件（确保文件存在且未被其他程序打开）
file_path = '1.xlsx'  # 替换为你的文件路径
workbook = openpyxl.load_workbook(file_path)

# 获取第一个工作表
sheet = workbook.worksheets[0]  # 索引0表示第一个sheet

# 在第七列（G列）写入556
column_index = 6  # Excel列号（A=1, B=2, ..., G=7）

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    cell = row[column_index - 1]  # Python索引从0开始，所以G列对应索引6
    cell.value = 556

# ❗直接保存到原文件（覆盖）
workbook.save(file_path)
print(f"已成功在第一个工作表的第{column_index}列写入数据，并保存到 {file_path}")
