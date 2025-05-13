import os
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment

def process_excel_file(source_file, target_file):
    """
    处理 XLSX 文件，删除第一列中"年月日"之后的字符串，增加所有单元格宽度，
    计算每个工作表第二列与第三列的总和差值，并将结果汇总到最后一个工作表中。
    
    参数:
        source_file (str): 源 XLSX 文件路径
        target_file (str): 目标 XLSX 文件路径
    """
    # 检查源文件是否存在
    if not os.path.exists(source_file):
        raise FileNotFoundError(f"源文件不存在: {source_file}")
    
    # 加载源工作簿
    wb = load_workbook(source_file)
    
    # 存储每个工作表的计算结果
    sheet_results = []
    
    # 处理每个工作表
    for sheet in wb.worksheets:
        # 获取最大行数
        max_row = sheet.max_row
        
        # 处理第一列的日期数据（删除"年月日"之后的字符串）
        for row_idx in range(1, max_row + 1):
            cell = sheet.cell(row=row_idx, column=1)
            if cell.value and isinstance(cell.value, str):
                # 使用正则表达式匹配并保留"年月日"格式
                match = re.search(r'(\d{4}年\d{1,2}月\d{1,2}日)', cell.value)
                if match:
                    cell.value = match.group(1)
        
        # 计算第二列和第三列的总和
        sum_col2 = 0
        sum_col3 = 0
        
        for row_idx in range(1, max_row + 1):
            # 第二列
            cell_col2 = sheet.cell(row=row_idx, column=2)
            if cell_col2.value and isinstance(cell_col2.value, (int, float)):
                sum_col2 += cell_col2.value
            
            # 第三列
            cell_col3 = sheet.cell(row=row_idx, column=3)
            if cell_col3.value and isinstance(cell_col3.value, (int, float)):
                sum_col3 += cell_col3.value
        
        # 计算差值并存储结果
        difference = sum_col2 - sum_col3
        sheet_results.append((sheet.title, difference))
        
        # 增加所有列宽（原宽度+5磅）
        max_col = sheet.max_column
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            # 获取原列宽（如果未设置则默认为8.43）
            original_width = sheet.column_dimensions[col_letter].width or 8.43
            # 增加5磅
            sheet.column_dimensions[col_letter].width = original_width + 5
    
    # 获取最后一个工作表（如果没有则创建）
    if not wb.worksheets:
        summary_sheet = wb.create_sheet("汇总")
    else:
        summary_sheet = wb.worksheets[-1]
    
    # 在汇总表中写入结果
    summary_sheet.cell(row=1, column=1).value = "工作表名称"
    summary_sheet.cell(row=1, column=2).value = "第二列减第三列的差值"
    
    # 写入每个工作表的结果
    for i, (sheet_name, result) in enumerate(sheet_results, start=2):
        summary_sheet.cell(row=i, column=1).value = sheet_name
        summary_sheet.cell(row=i, column=2).value = result
    
    # 计算所有sheet结果的汇总值并写入第二列第二排
    total_sum = sum(result for _, result in sheet_results)
    summary_sheet.cell(row=2, column=2).value = total_sum
    
    # 设置汇总表表头样式
    for col_idx in range(1, 3):
        cell = summary_sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 保存目标文件
    wb.save(target_file)
    print(f"✅ 已成功处理 {len(wb.worksheets)} 个工作表，并保存到 {target_file}")
    print(f"📊 汇总结果已写入最后一个工作表的第二列第二行: {total_sum}")

# 使用示例
if __name__ == "__main__":
    source_file = "zd3.xlsx"  # 替换为你的源文件路径
    target_file = "zd3.xlsx"  # 替换为你的目标文件路径
    
    try:
        process_excel_file(source_file, target_file)
    except Exception as e:
        print(f"❌ 处理过程中出错: {e}")
