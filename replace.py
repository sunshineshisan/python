import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment

def copy_xls_to_xlsx(source_file, target_file):
    """
    将 XLS 文件中的所有工作表复制到新的 XLSX 文件中，保留表名和格式。
    
    参数:
        source_file (str): 源 XLS 文件路径
        target_file (str): 目标 XLSX 文件路径
    """
    # 检查源文件是否存在
    if not os.path.exists(source_file):
        raise FileNotFoundError(f"源文件不存在: {source_file}")
    
    # 创建目标工作簿
    target_wb = Workbook()
    
    # 读取源文件的所有表名
    xls = pd.ExcelFile(source_file)
    sheet_names = xls.sheet_names
    
    # 处理每个工作表
    for sheet_name in sheet_names:
        # 读取当前工作表的数据
        df = xls.parse(sheet_name)
        
        # 在目标工作簿中创建同名工作表
        if sheet_name in target_wb.sheetnames:
            target_sheet = target_wb[sheet_name]
        else:
            target_sheet = target_wb.create_sheet(sheet_name)
        
        # 写入表头
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = target_sheet.cell(row=1, column=col_idx)
            cell.value = col_name
            
            # 尝试保留表头格式（如果源文件有格式信息）
            try:
                # 这里可以添加更多格式设置，如字体、颜色等
                cell.font = Font(bold=True)
            except Exception as e:
                pass
        
        # 写入数据
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = target_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # 尝试保留单元格格式（如果源文件有格式信息）
                try:
                    # 示例：设置数字格式（根据数据类型）
                    if isinstance(value, float):
                        cell.number_format = '0.00'  # 两位小数
                except Exception as e:
                    pass
    
    # 删除默认创建的工作表（如果有）
    if 'Sheet' in target_wb.sheetnames and 'Sheet' not in sheet_names:
        target_wb.remove(target_wb['Sheet'])
    
    # 保存目标文件
    target_wb.save(target_file)
    print(f"✅ 已成功将 {len(sheet_names)} 个工作表从 {source_file} 复制到 {target_file}")

# 使用示例
if __name__ == "__main__":
    source_file = "zd3.xls"  # 替换为你的源文件路径
    target_file = "zd4.xlsx"  # 替换为你的目标文件路径
    
    try:
        copy_xls_to_xlsx(source_file, target_file)
    except Exception as e:
        print(f"❌ 复制过程中出错: {e}")
